import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "output/screener_output.xlsx"


def export_to_excel(results):

    os.makedirs("output", exist_ok=True)

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        for preset_name, df in results.items():

            df = df.sort_values(
                by="composite_quality_score",
                ascending=False
            )

            df.to_excel(
                writer,
                sheet_name=preset_name[:31],
                index=False
            )

    workbook = load_workbook(OUTPUT_FILE)

    green = PatternFill(
        fill_type="solid",
        start_color="C6EFCE"
    )

    red = PatternFill(
        fill_type="solid",
        start_color="FFC7CE"
    )

    header = PatternFill(
        fill_type="solid",
        start_color="4F81BD"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    for sheet in workbook.worksheets:

        sheet.freeze_panes = "A2"

        # Header Formatting
        for cell in sheet[1]:

            cell.fill = header
            cell.font = white_font

        # Auto Width
        for column_cells in sheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value else 0
                for cell in column_cells
            )

            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 3

        # Conditional Formatting
        header_map = {}

        for cell in sheet[1]:
            header_map[cell.value] = cell.column

        if "composite_quality_score" in header_map:

            col = header_map["composite_quality_score"]

            for row in range(2, sheet.max_row + 1):

                value = sheet.cell(
                    row=row,
                    column=col
                ).value

                if value is None:
                    continue

                if value >= 70:
                    sheet.cell(
                        row=row,
                        column=col
                    ).fill = green

                else:
                    sheet.cell(
                        row=row,
                        column=col
                    ).fill = red

    workbook.save(OUTPUT_FILE)

    print(f"\nGenerated {OUTPUT_FILE}")