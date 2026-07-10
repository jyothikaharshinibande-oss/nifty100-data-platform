import os
import sqlite3
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

DB_PATH = "db/nifty100.db"
OUTPUT_FILE = "output/peer_comparison.xlsx"


class PeerComparisonReport:

    def __init__(self):

        self.conn = sqlite3.connect(DB_PATH)

        os.makedirs("output", exist_ok=True)

    def load_data(self):

        query = """

        SELECT

            pg.peer_group_name,

            pg.is_benchmark,

            c.company_name,

            fr.company_id,

            fr.year,

            fr.return_on_equity_pct,

            fr.operating_profit_margin_pct,

            fr.debt_to_equity,

            fr.interest_coverage,

            fr.asset_turnover,

            fr.free_cash_flow_cr,

            pp.metric,

            pp.percentile_rank

        FROM peer_groups pg

        INNER JOIN companies c
            ON pg.company_id = c.id

        INNER JOIN financial_ratios fr
            ON pg.company_id = fr.company_id

        LEFT JOIN peer_percentiles pp
            ON pp.company_id = fr.company_id

        """

        return pd.read_sql(query, self.conn)

    def export(self):

        df = self.load_data()

        if df.empty:

            print("No data found.")

            return

        with pd.ExcelWriter(
            OUTPUT_FILE,
            engine="openpyxl"
        ) as writer:

            for group in sorted(df.peer_group_name.unique()):

                sheet = df[
                    df.peer_group_name == group
                ]

                sheet.to_excel(
                    writer,
                    sheet_name=group[:31],
                    index=False
                )

        wb = load_workbook(OUTPUT_FILE)

        green = PatternFill(
            fill_type="solid",
            start_color="90EE90"
        )

        yellow = PatternFill(
            fill_type="solid",
            start_color="FFF59D"
        )

        red = PatternFill(
            fill_type="solid",
            start_color="FF9999"
        )

        gold = PatternFill(
            fill_type="solid",
            start_color="FFD700"
        )

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for ws in wb.worksheets:

            for cell in ws[1]:

                cell.fill = header_fill
                cell.font = header_font

            headers = [
                c.value
                for c in ws[1]
            ]

            if "percentile_rank" in headers:

                col = headers.index(
                    "percentile_rank"
                ) + 1

                for r in range(2, ws.max_row + 1):

                    value = ws.cell(
                        r,
                        col
                    ).value

                    if value is None:
                        continue

                    if value >= 0.75:

                        ws.cell(
                            r,
                            col
                        ).fill = green

                    elif value >= 0.25:

                        ws.cell(
                            r,
                            col
                        ).fill = yellow

                    else:

                        ws.cell(
                            r,
                            col
                        ).fill = red

            if "is_benchmark" in headers:

                bench = headers.index(
                    "is_benchmark"
                ) + 1

                for r in range(2, ws.max_row + 1):

                    if ws.cell(
                        r,
                        bench
                    ).value == 1:

                        for c in range(
                            1,
                            ws.max_column + 1
                        ):

                            ws.cell(
                                r,
                                c
                            ).fill = gold

            # Median Row

            last = ws.max_row + 1

            ws.cell(last, 1).value = "Median"

            numeric_cols = []

            for col in range(1, ws.max_column + 1):

                values = []

                for row in range(2, ws.max_row):

                    value = ws.cell(
                        row,
                        col
                    ).value

                    if isinstance(
                        value,
                        (int, float)
                    ):

                        values.append(value)

                if values:

                    ws.cell(
                        last,
                        col
                    ).value = float(
                        pd.Series(values).median()
                    )

        wb.save(OUTPUT_FILE)

        print(
            "Peer Comparison Report Generated"
        )

        self.conn.close()


if __name__ == "__main__":

    PeerComparisonReport().export()